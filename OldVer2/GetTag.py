from openpyxl import load_workbook
import re
import requests
import xlwt
import xlrd
from xlutils.copy import copy


def write_excel_xls(path, sheet_name, value):
    index = len(value)  # 获取需要写入数据的行数
    workbook = xlwt.Workbook()  # 新建一个工作簿
    sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])  # 像表格中写入数据（对应的行和列）
    workbook.save(path)  # 保存工作簿
    print("xls格式表格写入数据成功！")


def write_excel_xls_append(path, value):
    index = len(value)  # 获取需要写入数据的行数
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
    new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
    for i in range(0, index):
        for j in range(0, len(value[i])):
            new_worksheet.write(i + rows_old, j, value[i][j])  # 追加写入数据，注意是从i+rows_old行开始写入
    new_workbook.save(path)  # 保存工作簿
    print("xls格式表格【追加】写入数据成功！")


def read_excel_xls(path):
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    for i in range(0, worksheet.nrows):
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t", end="")  # 逐行逐列读取数据
        print()


def get_html_text(burl, self_header):
    try:
        response = requests.get(burl, headers=self_header, timeout=30)
        response.raise_for_status()
        # response.encoding = response.apparent_encoding
        response.encoding = 'utf-8'
        # print(response.text)
        return response.text
    except:
        return ""


def re_get_inf(html, rankArr, i):
    # tag_list = re.findall(r'<li class="tag">< a target= "_blank">([\s\S]*?)</ a><!----><!----></li>',html)
    # tags = re.findall(r'<a href="//search.bilibili.com/.*?"target="_blank">(.+?)</a>', html)
    tags = re.findall(r'target="_blank" class="tag-link">[<span>]*\s*([\u4e00-\u9fa5]*?)\s*[</span>]*</a>', html)
    tagsStr = ''
    tagsArr = []
    for tag in tags:
        print(tag)
        tagsStr = tagsStr + tag + ","
    tagsArr.append(i)  # 第i名的排名
    tagsArr.append(tagsStr[0:len(tagsStr) - 1])  # 第i名标签集
    rankArr.append(tagsArr)


####main函数
def GetTag():
    # 1.读取Excel元数据
    wb = load_workbook("bilibili_rankdata.xlsx")
    ws = wb.active
    rank = []
    for col in ws.iter_cols():
        rank.append(col)

    # 2.获取爬虫排行榜信息rankArr[]
    rankArr = []
    for i in range(1, len(rank[0])):
    # 获取排行榜第i名
        url = (rank[len(rank) - 1][i].value)
        self_header = {
             "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36"
        }
        # 获取第i名的链接html
        linkhtml = get_html_text(url, self_header)
        re_get_inf(linkhtml, rankArr, i)
        print('loading ', round((i / len(rank[0]) * 100), 2), '%', '=' * i)
        # if i == 5:
        #     break

    # 3.导出execl
    book_name_xls = 'bilibili_tag.xls'
    sheet_name_xls = '排行榜'
    value_title = [["id", "标签"], ]
    write_excel_xls(book_name_xls, sheet_name_xls, value_title)
    write_excel_xls_append(book_name_xls, rankArr)
    print("===========================ok=================================")
